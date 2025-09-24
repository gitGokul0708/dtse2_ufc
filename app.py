# @title Universal File-to-Text Converter
# This script is designed to run in a Google Colab environment.
# It provides functions to upload files, convert them to plain text or Markdown,
# and then display a preview and a download link.

import os
import io
import zipfile
import subprocess
from pathlib import Path
from google.colab import files
from IPython.display import display, HTML

# Quietly install required packages
print("Installing required packages...")
subprocess.run(['pip', 'install', 'openpyxl', 'pypandoc', 'beautifulsoup4', '--quiet'], capture_output=True, text=True)
subprocess.run(['apt-get', 'install', 'pandoc', '--quiet'], capture_output=True, text=True)
print("Installation complete.")

def upload_file():
  """
  Provides a file upload widget for the user in a Colab environment.
  Returns the file path of the uploaded file.
  """
  print("\nPlease upload your file(s) (Word, Excel, PPTX, HTML, ZIP, etc.).")
  uploaded = files.upload()
  if uploaded:
    # Get the file name from the uploaded dictionary
    file_name = list(uploaded.keys())[0]
    return file_name
  else:
    return None

def universal_file_converter(file_path):
  """
  Converts a wide variety of file types into a single Markdown/text string.

  Args:
    file_path (str): The path to the file to be converted.

  Returns:
    tuple: A tuple containing the converted text (str) and a message (str).
  """
  # Determine the file type based on its extension
  file_extension = Path(file_path).suffix.lower()

  # Output placeholder
  output_text = ""
  message = f"Successfully converted {file_path} to Markdown."

  try:
    if file_extension in ['.docx', '.pptx', '.odt', '.rtf']:
      # Use pypandoc for direct document conversion
      import pypandoc
      output_text = pypandoc.convert_file(file_path, to='markdown_strict')
    elif file_extension in ['.xlsx']:
      # Handle Excel files using openpyxl
      from openpyxl import load_workbook
      workbook = load_workbook(file_path)
      for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        output_text += f"\n\n# Excel Sheet: {sheet_name}\n"
        for row in sheet.iter_rows():
          row_text = ' | '.join([str(cell.value) if cell.value is not None else '' for cell in row])
          output_text += row_text + '\n'
    elif file_extension in ['.html', '.htm']:
      # Use BeautifulSoup for HTML files
      from bs4 import BeautifulSoup
      with open(file_path, 'r', encoding='utf-8') as f:
        html_content = f.read()
      soup = BeautifulSoup(html_content, 'html.parser')
      # Extract all text, clean up extra spaces and newlines
      text_content = soup.get_text()
      output_text = os.linesep.join([s for s in text_content.splitlines() if s])
    elif file_extension == '.zip':
      # Handle zip files by extracting and converting each file
      with zipfile.ZipFile(file_path, 'r') as zip_ref:
        for member in zip_ref.infolist():
          if not member.is_dir():
            extracted_path = zip_ref.extract(member)
            converted_content, _ = universal_file_converter(extracted_path)
            output_text += f"\n\n-- ZIP Archive: {member.filename} --\n"
            output_text += converted_content
            os.remove(extracted_path) # Clean up extracted file
    elif file_extension == '.pdf':
      message = "PDF conversion requires additional, complex packages. This function does not support it."
      output_text = ""
    else:
      # For all other text-based files, read them directly
      with open(file_path, 'r', encoding='utf-8') as f:
        output_text = f.read()

  except Exception as e:
    message = f"An error occurred during conversion: {e}"
    output_text = ""

  return output_text, message

def display_and_download_output(text, file_name):
  """
  Displays a preview of the converted text and provides a download link.

  Args:
    text (str): The converted text to display.
    file_name (str): The original file name used for the download link.
  """
  if not text:
    print("No content to display or download.")
    return

  # Display the first 1000 characters as a preview
  preview_text = text[:1000] + ('...' if len(text) > 1000 else '')
  print("\n--- Preview (first 1000 characters) ---")
  print(preview_text)
  print("--- End of Preview ---")

  # Create a download link for the full text
  download_file_name = f"{Path(file_name).stem}_converted.md"
  with open(download_file_name, 'w', encoding='utf-8') as f:
    f.write(text)

  # Use google.colab.files.download to download the file directly
  files.download(download_file_name)


# --- Main execution block ---
if __name__ == "__main__":
  uploaded_file_path = upload_file()
  if uploaded_file_path:
    converted_text, status_message = universal_file_converter(uploaded_file_path)
    print(status_message)
    display_and_download_output(converted_text, uploaded_file_path)
  else:
    print("No file was uploaded. Please try again.")
